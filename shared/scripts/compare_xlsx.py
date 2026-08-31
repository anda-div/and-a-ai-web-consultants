# -*- coding: utf-8 -*-
r"""2つの xlsx をセル単位で突き合わせる（取得方式を切り替える前の検証用）

これまでの方法（GAS・管理画面からの書き出し）で作った Excel と、
この PC から取った xlsx が**同じ値になっているか**を1セルずつ確かめる。
**合計が合っているだけでは足りない。**

ある案件では 5,714セル中4セルだけが丸め方の違いで食い違った。
合計は完全に一致していたため、セル単位で比べなければ気づけなかった。

    python compare_xlsx.py 旧.xlsx 新.xlsx
    python compare_xlsx.py 旧.xlsx 新.xlsx --sheet 01_KPIサマリ
    python compare_xlsx.py 旧.xlsx 新.xlsx --tolerance 0.05

出るもの
  ・シートごとの 一致 / 不一致 / 行数の差
  ・不一致は「どのシートの何行目・何列目が、いくつ対いくつか」まで出す

終了コードは、1件でも不一致があれば 1。
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys


def norm(v):
    """比較のために値をそろえる。

    Excel は日付らしい文字列を日付に変換してしまう（'2026-01' が
    2026-01-01 になる）。表示上は同じなので、文字列に戻して比べる。
    """
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return f"{v.year:04d}-{v.month:02d}"
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, (int, float)):
        f = float(v)
        return int(f) if f.is_integer() else round(f, 4)
    return str(v)


def close(a, b, tol) -> bool:
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if a == b:
            return True
        base = max(abs(float(a)), abs(float(b)), 1.0)
        return abs(float(a) - float(b)) / base <= tol
    return a == b


def rows_of(ws) -> list[list]:
    out = []
    for r in ws.values:
        if r is None:
            continue
        cells = [norm(c) for c in r]
        if any(c != "" for c in cells):
            out.append(cells)
    return out


def col_name(i: int) -> str:
    s = ""
    i += 1
    while i:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    ap = argparse.ArgumentParser(description="2つの xlsx をセル単位で突き合わせる")
    ap.add_argument("old", help="基準にする方（GASで取ったもの）")
    ap.add_argument("new", help="確かめる方（この PC で取ったもの）")
    ap.add_argument("--sheet", help="1シートだけ比べる")
    ap.add_argument("--tolerance", type=float, default=0.0,
                    help="数値の許容誤差（割合）。既定 0＝完全一致を求める")
    ap.add_argument("--max-show", type=int, default=12, help="1シートあたりの表示件数")
    a = ap.parse_args()

    import openpyxl
    old = openpyxl.load_workbook(a.old, data_only=True)
    new = openpyxl.load_workbook(a.new, data_only=True)

    def sheets(wb):
        return [s for s in wb.sheetnames
                if any(any(c is not None and str(c).strip() for c in (r or []))
                       for r in wb[s].values)]

    o_sheets, n_sheets = sheets(old), sheets(new)
    targets = [a.sheet] if a.sheet else [s for s in n_sheets if s in o_sheets]

    only_old = [s for s in o_sheets if s not in n_sheets]
    only_new = [s for s in n_sheets if s not in o_sheets]
    if only_old:
        print(f"※ 旧にだけあるシート: {', '.join(only_old)}")
    if only_new:
        print(f"※ 新にだけあるシート: {', '.join(only_new)}")
    if only_old or only_new:
        print()

    total_cells = total_diff = 0
    bad_sheets = []

    for name in targets:
        o, n = rows_of(old[name]), rows_of(new[name])
        diffs = []
        for ri in range(max(len(o), len(n))):
            orow = o[ri] if ri < len(o) else []
            nrow = n[ri] if ri < len(n) else []
            for ci in range(max(len(orow), len(nrow))):
                ov = orow[ci] if ci < len(orow) else ""
                nv = nrow[ci] if ci < len(nrow) else ""
                total_cells += 1
                if not close(ov, nv, a.tolerance):
                    diffs.append((ri + 1, ci, ov, nv))
        total_diff += len(diffs)
        mark = "一致" if not diffs and len(o) == len(n) else "不一致"
        rowinfo = f"{len(o)}行" if len(o) == len(n) else f"{len(o)}行 → {len(n)}行"
        print(f"{'OK ' if mark == '一致' else 'NG '} {name:<24} {rowinfo:<16} "
              f"差 {len(diffs)} セル")
        if diffs:
            bad_sheets.append(name)
            for ri, ci, ov, nv in diffs[:a.max_show]:
                print(f"      {col_name(ci)}{ri}  旧「{ov}」 → 新「{nv}」")
            if len(diffs) > a.max_show:
                print(f"      …ほか {len(diffs) - a.max_show} セル")

    print()
    print(f"比べたセル {total_cells:,} / 差のあったセル {total_diff:,}")
    if total_diff == 0:
        print("完全に一致しました。取得方式を切り替えても、下流は影響を受けません。")
        return 0
    print(f"不一致のシート: {', '.join(bad_sheets)}")
    print("原因はたいてい次のどちらかです。")
    print("  ・取得の定義の違い（指標・絞り込み・並び順・丸め方・表記）→ こちらを直す")
    print("  ・元データ側の再処理（同じ日に両方を取り直して比べると切り分けられます）")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
